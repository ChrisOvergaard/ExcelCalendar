import datetime

from openpyxl import *
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

year = 2026

wb = Workbook()
sheet = wb.active

grey_fill = PatternFill(start_color='DDDDDD', fill_type='solid')

def date_format(d: datetime.datetime):
    match d.weekday():
        case 0:
            return "M " + str(d.day)
        case 1:
            return "T " + str(d.day)
        case 2:
            return "O " + str(d.day)
        case 3:
            return "T " + str(d.day)
        case 4:
            return "F " + str(d.day)
        case 5:
            return "L " + str(d.day)
        case 6:
            return "S " + str(d.day)

for col in range(12):
    date = datetime.datetime(year, col+1, 1)
    monthCell = sheet.cell(1, col * 3+2, date.strftime("%B " + str(year)))
    monthCell.alignment = Alignment(horizontal='center')
    monthCell.font = Font(name='Trebuchet MS', size=9, bold=True)



    currentMonth = date.month
    activeRow = 2
    while date.month == currentMonth:
        dateCell = sheet.cell(activeRow,col * 3+1, date_format(date))

        sheet.column_dimensions[get_column_letter(col * 3 + 1)].width = 5
        sheet.column_dimensions[get_column_letter(col * 3 + 2)].width = 20
        sheet.column_dimensions[get_column_letter(col * 3 + 3)].width = 5

        if date.weekday() == 0:
            weekNoCell = sheet.cell(activeRow, col * 3+3, date.strftime("%V"))
        else:
            weekNoCell = sheet.cell(activeRow, col * 3 + 3, "")

        if date.weekday() >= 5:
            dateCell.fill = grey_fill
        if date.weekday() >= 6:
            sheet.cell(activeRow,  col * 3+2).fill = grey_fill
            sheet.cell(activeRow, col * 3 + 3).fill = grey_fill

        date = date + datetime.timedelta(days=1)
        activeRow += 1


wb.save("Sweeper Calendar " + str(year) + ".xlsx")